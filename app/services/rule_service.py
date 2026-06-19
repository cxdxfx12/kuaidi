"""
运费规则管理服务
支持可视化配置、保存、读取
规则存储为 JSON 文件（data/config/fee_rules.json）
三层规则体系：客户级 → 区域级 → 全局级
"""
import os
import json
import math
from typing import Dict, List, Optional

# 用于 Excel 解析/写入
import xlsxwriter
try:
    import openpyxl
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


def apply_weight_rounding(weight: float, mode: str, params: Optional[Dict] = None) -> float:
    """
    根据重量进位模式调整重量
    
    :param weight: 原始重量（kg）
    :param mode: 进位模式:
        - "actual": 实际重量
        - "round_05": 0.5进位（<0.5取0.5, ≥0.5取1）
        - "round_1": 四舍五入（<0.5舍去, ≥0.5取1）
        - "ceil_1kg": 向上取整（小数位都向上取整）
        - "segmented": 分段进位（自定义舍位/进位值）
        - "round_trunc": 进位舍位（只舍或只进）
    :param params: 进位参数:
        - segmented: {"segment_drop": 0.2, "segment_ceil": 0.7}
        - round_trunc: {"direction": "drop" | "ceil"}
    :return: 调整后的重量
    """
    params = params or {}
    
    if mode == "actual":
        return weight
    elif mode == "round_05":
        if weight <= 0.5:
            return 0.5
        return math.ceil(weight) if weight > 1.0 else 1.0
    elif mode == "round_1":
        return round(weight)
    elif mode == "ceil_1kg":
        return math.ceil(weight)
    elif mode == "segmented":
        segment_drop = params.get("segment_drop", 0.2)
        segment_ceil = params.get("segment_ceil", 0.7)
        if weight <= segment_drop:
            return 0.0
        elif weight <= segment_ceil:
            return (segment_drop + segment_ceil) / 2
        else:
            return math.ceil(weight) if weight > 1.0 else 1.0
    elif mode == "round_trunc":
        direction = params.get("direction", "drop")
        if direction == "drop":
            return float(int(weight))
        else:
            return math.ceil(weight)
    return weight


class Rule:
    """单条计费规则"""
    def __init__(self, name: str = "", regions: str = "", stations: str = "", 
                 min_weight: float = 0.0, max_weight: float = 999.0, 
                 first_fee: float = 0.0, continued_fee: float = 0.0, 
                 min_fee: float = 0.0, rule_type: str = "region",
                 continued_unit: str = "kg", weight_rounding: str = "actual",
                 rounding_params: Optional[Dict] = None):
        """
        :param name: 规则名称
        :param regions: 逗号分隔的区域关键词，如 "上海,江苏,浙江"
        :param stations: 逗号分隔的网点编码，如 "ST001,ST002"
        :param rule_type: 规则类型: "station"(网点级), "region"(区域级), "global"(全局级)
        :param continued_unit: 续重单位: "kg"(全续), "100g"(百克续)
        :param weight_rounding: 重量进位模式: "actual", "round_05", "round_1", "ceil_1kg", "segmented", "round_trunc"
        :param rounding_params: 进位参数: {"segment_drop": 0.2, "segment_ceil": 0.7, "direction": "drop"}
        """
        self.name = name
        self.regions = regions
        self.stations = stations
        self.min_weight = float(min_weight)
        self.max_weight = float(max_weight)
        self.first_fee = float(first_fee)
        self.continued_fee = float(continued_fee)
        self.min_fee = float(min_fee)
        self.rule_type = rule_type  # "station", "region", "global"
        self.continued_unit = continued_unit  # "kg" or "100g"
        self.weight_rounding = weight_rounding  # 重量进位模式
        self.rounding_params = rounding_params or {}  # 进位参数

    def matches(self, region: str = "", station_code: str = "", weight: float = 0.0) -> bool:
        """判断是否匹配该规则"""
        # 网点匹配
        if self.stations.strip():
            station_list = [s.strip() for s in self.stations.split(",") if s.strip()]
            if station_code and station_code.strip() not in station_list:
                return False
        
        # 区域匹配（包含任一关键词即可）
        region_match = False
        region = (region or "").strip()
        if self.regions.strip():
            keywords = [k.strip() for k in self.regions.split(",") if k.strip()]
            region_match = any(k in region for k in keywords)
        else:
            region_match = True

        # 重量匹配
        weight_match = self.min_weight <= float(weight) <= self.max_weight

        return region_match and weight_match

    def to_dict(self) -> Dict:
        return {
            "name": self.name,
            "regions": self.regions,
            "stations": self.stations,
            "min_weight": self.min_weight,
            "max_weight": self.max_weight,
            "first_fee": self.first_fee,
            "continued_fee": self.continued_fee,
            "min_fee": self.min_fee,
            "rule_type": self.rule_type,
            "continued_unit": self.continued_unit,
            "weight_rounding": self.weight_rounding,
            "rounding_params": self.rounding_params,
        }

    @classmethod
    def from_dict(cls, data: Dict) -> "Rule":
        return cls(
            name=data.get("name", ""),
            regions=data.get("regions", ""),
            stations=data.get("stations", ""),
            min_weight=float(data.get("min_weight", 0)),
            max_weight=float(data.get("max_weight", 999)),
            first_fee=float(data.get("first_fee", 0)),
            continued_fee=float(data.get("continued_fee", 0)),
            min_fee=float(data.get("min_fee", 0)),
            rule_type=data.get("rule_type", "region"),
            continued_unit=data.get("continued_unit", "kg"),
            weight_rounding=data.get("weight_rounding", "actual"),
            rounding_params=data.get("rounding_params", {}),
        )


class RuleService:
    """规则管理服务 - 支持三层规则体系"""

    def __init__(self):
        from app.models.path_config import get_config_file
        self.config_file = get_config_file("fee_rules.json")
        self.default_settings_file = get_config_file("default_settings.json")
        # 加载"无重量默认价格"（取不到则用3.0兜底）
        self._empty_weight_fee = self._load_empty_weight_fee()

    def _load_empty_weight_fee(self) -> float:
        try:
            if os.path.exists(self.default_settings_file):
                with open(self.default_settings_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                val = data.get("empty_weight_fee")
                if val is not None:
                    return float(val)
        except Exception:
            pass
        return 3.0

    def load_rules(self) -> List[Rule]:
        """读取规则列表（按类型分组）"""
        if not os.path.exists(self.config_file):
            rules = self._create_default_rules()
            self.save_rules(rules)
            return rules

        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return [Rule.from_dict(r) for r in data.get("rules", [])]
        except Exception as e:
            print(f"读取规则失败: {e}")
            return self._create_default_rules()

    def _create_default_rules(self) -> List[Rule]:
        """创建默认规则（包含网点级示例规则）"""
        rules = []
        
        # 全局兜底规则（优先级最低）
        rules.append(Rule("全局规则", "", "", 0, 999, 6.0, 3.0, 6.0, "global"))
        
        # 区域级规则
        region_rules = [
            ("北京 - 华北地区", "北京", 4.0, 1.8, 4.0),
            ("天津 - 华北地区", "天津", 4.0, 1.8, 4.0),
            ("上海 - 华东地区", "上海", 3.5, 1.5, 3.5),
            ("重庆 - 西南地区", "重庆", 5.0, 2.5, 5.0),
            ("江苏 - 华东地区", "江苏", 3.5, 1.5, 3.5),
            ("浙江 - 华东地区", "浙江", 3.5, 1.5, 3.5),
            ("安徽 - 华东地区", "安徽", 3.5, 1.5, 3.5),
            ("福建 - 华东地区", "福建", 3.5, 1.5, 3.5),
            ("江西 - 华东地区", "江西", 3.5, 1.5, 3.5),
            ("山东 - 华东地区", "山东", 3.5, 1.5, 3.5),
            ("河北 - 华北地区", "河北", 4.0, 1.8, 4.0),
            ("山西 - 华北地区", "山西", 4.0, 1.8, 4.0),
            ("内蒙古 - 华北地区", "内蒙古", 4.0, 1.8, 4.0),
            ("广东 - 华南地区", "广东", 4.0, 1.8, 4.0),
            ("广西 - 华南地区", "广西", 4.0, 1.8, 4.0),
            ("海南 - 华南地区", "海南", 4.0, 1.8, 4.0),
            ("河南 - 华中地区", "河南", 4.5, 2.0, 4.5),
            ("湖北 - 华中地区", "湖北", 4.5, 2.0, 4.5),
            ("湖南 - 华中地区", "湖南", 4.5, 2.0, 4.5),
            ("四川 - 西南地区", "四川", 5.0, 2.5, 5.0),
            ("贵州 - 西南地区", "贵州", 5.0, 2.5, 5.0),
            ("云南 - 西南地区", "云南", 5.0, 2.5, 5.0),
            ("西藏 - 西南地区", "西藏", 5.0, 2.5, 5.0),
            ("陕西 - 西北地区", "陕西", 8.0, 4.0, 8.0),
            ("甘肃 - 西北地区", "甘肃", 8.0, 4.0, 8.0),
            ("青海 - 西北地区", "青海", 8.0, 4.0, 8.0),
            ("宁夏 - 西北地区", "宁夏", 8.0, 4.0, 8.0),
            ("新疆 - 西北地区", "新疆", 8.0, 4.0, 8.0),
            ("辽宁 - 东北地区", "辽宁", 5.0, 2.5, 5.0),
            ("吉林 - 东北地区", "吉林", 5.0, 2.5, 5.0),
            ("黑龙江 - 东北地区", "黑龙江", 5.0, 2.5, 5.0),
            ("香港 - 港澳台", "香港", 30.0, 20.0, 30.0),
            ("澳门 - 港澳台", "澳门", 30.0, 20.0, 30.0),
            ("台湾 - 港澳台", "台湾", 30.0, 20.0, 30.0),
        ]
        
        for name, regions, first_fee, continued_fee, min_fee in region_rules:
            rules.append(Rule(name, regions, "", 0, 999, first_fee, continued_fee, min_fee, "region"))
        
        # 网点级示例规则（优先级最高）
        rules.append(Rule("上海总部网点", "", "ST001", 0, 999, 3.0, 1.2, 3.0, "station"))
        rules.append(Rule("杭州旗舰网点", "", "ST002", 0, 999, 3.2, 1.3, 3.2, "station"))
        rules.append(Rule("北京核心网点", "", "ST003", 0, 999, 3.8, 1.6, 3.8, "station"))
        
        return rules

    def save_rules(self, rules: List[Rule], promotion_rules: Optional[List[Dict]] = None) -> bool:
        """保存规则列表（支持同时保存活动加价规则）"""
        try:
            # 清理旧版" - 华东地区/华北/西北"等大区后缀的默认规则（避免污染当前规则）
            cleaned_rules = []
            for r in rules:
                rname = r.name or ""
                # 旧默认区域名格式："省份 - 华东地区" / "省份 - 华北地区" / "省份 - 东北地区" / "省份 - 西北地区" / "省份 - 华南地区" / "省份 - 华中地区" / "省份 - 西南地区" / "省份 - 港澳台"
                legacy_suffixes = (
                    " - 华东地区", " - 华北地区", " - 东北地区", " - 西北地区",
                    " - 华南地区", " - 华中地区", " - 西南地区", " - 港澳台"
                )
                if any(rname.endswith(suf) for suf in legacy_suffixes):
                    # 跳过
                    continue
                cleaned_rules.append(r)

            data = {
                "version": "2.1",
                "description": "大圣运费计算规则 - 支持三层规则体系（客户级→区域级→全局级）+ 活动加价",
                "updated_at": __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "rules": [r.to_dict() for r in cleaned_rules],
                "promotion_rules": promotion_rules or []
            }
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存规则失败: {e}")
            return False

    # ============ Excel 导入/导出（批量导入客户+专属规则 ============

    PROVINCE_GROUP_MAP = {
        # key: group name in UI, value: list of province names
        "一区": ["浙江"],
        "二区": ["江苏", "安徽"],
        "三区": ["天津", "河北", "山东", "山西", "河南"],
        "四区": ["上海"],
        "五区": ["北京"],
        "六区": ["重庆"],
        "七区": ["广东", "广西", "海南", "福建", "江西", "湖南", "湖北"],
        "八区": ["黑龙江", "吉林", "辽宁", "内蒙古"],
        "九区": ["四川", "贵州", "云南"],
        "十区": ["甘肃", "宁夏", "青海", "陕西"],
        "十一区": ["新疆"],
        "十二区": ["西藏"],
        "十三区": ["海南"],  # 注：海南在原结构中存在
        "十四区": ["香港", "澳门", "台湾"],
    }

    # 反向映射：省份 → 组名
    _PROVINCE_TO_GROUP = {}
    for group, provinces in PROVINCE_GROUP_MAP.items():
        for p in provinces:
            _PROVINCE_TO_GROUP[p] = group
            _PROVINCE_TO_GROUP[p] = _PROVINCE_TO_GROUP.get(p, group)
            _PROVINCE_TO_GROUP[p] = _PROVINCE_TO_GROUP.get(p, group)
    del p
    del group
    del provinces

    WEIGHT_ROUNDING_TEXT_TO_MODE = {
        "实际重量": "actual",
        "0.5进位": "round_05",
        "四舍五入": "round_1",
        "向上取整": "ceil_1kg",
        "分段进位": "segmented",
        "进位舍位": "round_trunc",
    }
    CONTINUED_UNIT_TEXT_TO_MODE = {
        "全续": "kg",
        "百克续": "100g",
    }

    def generate_import_template(self, output_path: str) -> bool:
        """生成 Excel 导入模板（两个 Sheet：客户档案 + 客户专属规则）

        :param output_path: 输出路径（含 .xlsx 结尾）
        :return: 是否成功
        """
        try:
            wb = xlsxwriter.Workbook(output_path)

            # ---------- Sheet 1: 客户档案 ----------
            ws1 = wb.add_worksheet("客户档案")

            header_fmt = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})
            required_fmt = wb.add_format({"italic": True, "color": "#C00000", "border": 1})
            normal_fmt = wb.add_format({"border": 1})

            headers_station = ["客户编码*", "客户名称*", "联系人", "联系电话", "地址", "是否启用", "备注"]
            for col, header in enumerate(headers_station):
                ws1.write(0, col, header, header_fmt)

            # 示例行：3 行
            sample_stations = [
                ["C001", "蜜丝婷", "张经理", "13800138000", "上海市浦东新区XX路XX号", 1, "大客户，每月结算"],
                ["C002", "珀莱雅", "李经理", "13800138001", "杭州市西湖区XX路XX号", 1, "大客户"],
                ["C010", "淘宝店铺A", "", "", "", 1, ""],
            ]
            for row_idx, row_data in enumerate(sample_stations):
                for col, value in enumerate(row_data):
                    ws1.write(row_idx + 1, col, value, normal_fmt)

            # 列宽
            ws1.set_column("A:A", 12)  # 客户编码
            ws1.set_column("B:B", 20)  # 客户名称
            ws1.set_column("C:C", 12)  # 联系人
            ws1.set_column("D:D", 16)  # 联系电话
            ws1.set_column("E:E", 40)  # 地址
            ws1.set_column("F:F", 10)  # 是否启用
            ws1.set_column("G:G", 30)  # 备注

            # ---------- Sheet 2: 客户专属规则 ----------
            ws2 = wb.add_worksheet("客户专属规则")

            headers_rule = [
                "客户编码*",
                "分区名称*",
                "首重费(元)*",
                "续重费(元)*",
                "保底费(元)*",
                "续重单位",
                "重量进位",
            ]
            for col, header in enumerate(headers_rule):
                ws2.write(0, col, header, header_fmt)

            # 示例行：C001 的 3 个分区规则
            sample_rules = [
                ["C001", "一区", 3.5, 1.5, 2.0, "全续", "实际重量"],
                ["C001", "二区", 3.5, 1.5, 2.0, "全续", "实际重量"],
                ["C001", "三区", 4.0, 2.0, 3.0, "全续", "实际重量"],
                ["C001", "四区", 3.5, 1.5, 3.5, "全续", "实际重量"],
                ["C001", "五区", 4.0, 2.0, 3.0, "全续", "实际重量"],
                ["C001", "六区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C001", "七区", 4.5, 2.0, 4.5, "全续", "实际重量"],
                ["C001", "八区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C001", "九区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C001", "十区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C001", "十一区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C001", "十二区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C001", "十三区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C001", "十四区", 30.0, 20.0, 30.0, "全续", "实际重量"],
                # 2: 珀莱雅 - 江苏
                ["C002", "一区", 3.5, 1.5, 2.0, "全续", "实际重量"],
                ["C002", "二区", 3.5, 1.5, 2.0, "全续", "实际重量"],
                ["C002", "三区", 4.0, 2.0, 3.0, "全续", "实际重量"],
                ["C002", "四区", 3.5, 1.5, 3.5, "全续", "实际重量"],
                ["C002", "五区", 4.0, 2.0, 3.0, "全续", "实际重量"],
                ["C002", "六区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C002", "七区", 4.5, 2.0, 4.5, "全续", "实际重量"],
                ["C002", "八区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C002", "九区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C002", "十区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C002", "八区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C002", "九区", 5.0, 2.5, 5.0, "全续", "实际重量"],
                ["C002", "十区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C002", "十一区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C002", "十二区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C002", "十三区", 8.0, 4.0, 8.0, "全续", "实际重量"],
                ["C002", "十四区", 30.0, 20.0, 30.0, "全续", "实际重量"],
                # C010: 淘宝店铺A - 5 个分区
                ["C010", "一区", 4.0, 2.0, 2.0, "全续", "实际重量"],
                ["C010", "二区", 4.0, 2.0, 2.0, "全续", "实际重量"],
            ]

            for row_idx, row_data in enumerate(sample_rules):
                for col, value in enumerate(row_data):
                    ws2.write(row_idx + 1, col, value, normal_fmt)

            # 列宽
            ws2.set_column("A:A", 12)  # 客户编码
            ws2.set_column("B:B", 10)  # 分区名称
            ws2.set_column("C:C", 12)  # 首重费
            ws2.set_column("D:D", 12)  # 续重费
            ws2.set_column("E:E", 12)  # 保底费
            ws2.set_column("F:F", 10)  # 续重单位
            ws2.set_column("G:G", 12)  # 重量进位

            wb.close()
            return True
        except Exception as e:
            return False

    def parse_import_excel(self, file_path: str) -> Dict:
        """解析导入 Excel 文件（使用 openpyxl）
        :return: {"stations": [...], "rules": [...], "errors": [...], "warnings": [...]}
        """
        try:
            result = {"stations": [], "rules": [], "errors": [], "warnings": []}

            if not _HAS_OPENPYXL:
                result["errors"].append("缺少 openpyxl 依赖包，请安装: pip install openpyxl")
                return result

            wb = openpyxl.load_workbook(file_path, data_only=True)

            # ---------- Sheet 1: 客户档案 ----------
            try:
                if "客户档案" not in wb.sheetnames:
                    result["warnings"].append("找不到 [客户档案] Sheet，跳过客户档案导入")
                else:
                    ws_stations = wb["客户档案"]
                    for row_idx, row in enumerate(ws_stations.iter_rows(min_row=2, values_only=True)):
                        if not row or row[0] is None or str(row[0]).strip() == "":
                            continue
                        code = str(row[0]).strip()
                        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                        contact = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                        phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                        address = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                        is_active_val = row[5] if len(row) > 5 and row[5] is not None else 1
                        try:
                            is_active = bool(int(float(is_active_val)))
                        except Exception:
                            is_active = True
                        remark = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

                        if not name:
                            result["errors"].append(f"客户档案第 {row_idx + 2} 行（{code}）：客户名称不能为空")
                            continue
                        result["stations"].append({
                            "code": code, "name": name, "contact": contact,
                            "phone": phone, "address": address, "is_active": is_active,
                            "remark": remark, "row_num": row_idx + 2,
                        })
            except Exception as e:
                result["errors"].append(f"读取 [客户档案] Sheet 失败: {e}")

            # ---------- Sheet 2: 客户专属规则 ----------
            try:
                if "客户专属规则" not in wb.sheetnames:
                    result["warnings"].append("找不到 [客户专属规则] Sheet，跳过规则导入")
                else:
                    ws_rules = wb["客户专属规则"]
                    for row_idx, row in enumerate(ws_rules.iter_rows(min_row=2, values_only=True)):
                        if not row or row[0] is None or str(row[0]).strip() == "":
                            continue
                        if len(row) < 5:
                            continue

                        code = str(row[0]).strip()
                        group = str(row[1]).strip() if row[1] is not None else ""
                        first_fee_raw = row[2]
                        continued_fee_raw = row[3] if len(row) > 3 else 0
                        min_fee_raw = row[4] if len(row) > 4 else 0
                        continued_unit_text = str(row[5]).strip() if len(row) > 5 and row[5] is not None else "全续"
                        weight_rounding_text = str(row[6]).strip() if len(row) > 6 and row[6] is not None else "实际重量"

                        if not group:
                            result["errors"].append(f"规则第 {row_idx + 2} 行：分区名称不能为空")
                            continue

                        try:
                            first_fee = float(first_fee_raw) if first_fee_raw is not None else 0.0
                            continued_fee = float(continued_fee_raw) if continued_fee_raw is not None else 0.0
                            min_fee = float(min_fee_raw) if min_fee_raw is not None else 0.0
                        except (ValueError, TypeError):
                            result["errors"].append(f"规则第 {row_idx + 2} 行（{code}/{group}）：首重费/续重费/保底费必须是数字")
                            continue

                        continued_unit = self.CONTINUED_UNIT_TEXT_TO_MODE.get(continued_unit_text, "kg")
                        weight_rounding = self.WEIGHT_ROUNDING_TEXT_TO_MODE.get(weight_rounding_text, "actual")

                        provinces = self._resolve_provinces_from_group(group)
                        if not provinces:
                            result["errors"].append(f"规则第 {row_idx + 2} 行（{code}/{group}）：未知的分区名或省份名")
                            continue

                        for province in provinces:
                            result["rules"].append({
                                "code": code, "province": province, "group": group,
                                "first_fee": first_fee, "continued_fee": continued_fee, "min_fee": min_fee,
                                "continued_unit": continued_unit, "weight_rounding": weight_rounding,
                                "row_num": row_idx + 2,
                            })
            except Exception as e:
                result["errors"].append(f"读取 [客户专属规则] Sheet 失败: {e}")

            wb.close()
            return result
        except Exception as e:
            return {
                "stations": [], "rules": [],
                "errors": [f"解析 Excel 文件失败: {e}"],
                "warnings": [],
            }

    def _resolve_provinces_from_group(self, group_text: str) -> List[str]:
        """从"分区名称"解析为实际省份列表
        :param group_text: 分区名（如"一区"、"北京"等
        :return: 省份列表，空 list 表示不合法
        """
        if not group_text:
            return []
        gt = group_text.strip()
        if gt in self.PROVINCE_GROUP_MAP:
            return list(self.PROVINCE_GROUP_MAP[gt])
        # 如果直接传入的是单个省份
        if gt in self._PROVINCE_TO_GROUP:
            return [gt]
        # 支持多省份，用逗号分隔？目前不支持，但简单判断
        return []

    def save_import_result(self, parsed_result: Dict, conflict_mode: str = "skip") -> Dict:
        """将解析结果写入数据库和 JSON 文件

        :param parsed_result: parse_import_excel() 的返回值
        :param conflict_mode: "overwrite"（覆盖）| "skip"（跳过已存在客户） | "append"（追加规则）
        :return: {"success": bool, "message": str, "stats": {...}}
        """
        try:
            stations = parsed_result.get("stations", [])
            rules = parsed_result.get("rules", [])

            if not stations and not rules:
                return {"success": False, "message": "Excel 文件为空或未解析出有效数据",
                        "stats": {"inserted_customers": 0, "inserted_rules": 0}}

            from app.models.database import get_session
            from app.models.station import Station
            session = get_session()

            stats = {
                "inserted_customers": 0, "updated_customers": 0, "skipped_customers": 0,
                "inserted_rules": 0, "updated_rules": 0, "skipped_rules": 0,
            }

            try:
                # ========= 1. 处理客户档案 =========
                existing_codes = {s.station_code for s in session.query(Station).all()}
                for s_data in stations:
                    code = s_data["code"]
                    if code in existing_codes:
                        if conflict_mode == "overwrite":
                            station = session.query(Station).filter(Station.station_code == code).first()
                            if station:
                                station.station_name = s_data["name"]
                                station.address = s_data.get("address", "")
                                station.contact_person = s_data.get("contact", "")
                                station.contact_phone = s_data.get("phone", "")
                                station.is_active = s_data.get("is_active", True)
                                stats["updated_customers"] += 1
                        else:
                            stats["skipped_customers"] += 1
                    else:
                        station = Station(station_code=code, station_name=s_data["name"],
                                          address=s_data.get("address", ""),
                                          contact_person=s_data.get("contact", ""),
                                          contact_phone=s_data.get("phone", ""),
                                          is_active=s_data.get("is_active", True))
                        session.add(station)
                        stats["inserted_customers"] += 1
                session.commit()

                # ========= 2. 处理客户规则 =========
                all_rules = self.load_rules()
                existing_rule_keys = set()
                for r in all_rules:
                    if r.rule_type == "station" and r.stations and r.stations.strip():
                        station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                        province_parts = (r.regions or "").split(",")
                        province_list = [p.strip() for p in province_parts if p and p.strip()]
                        for sc in station_list:
                            for pr in province_list:
                                existing_rule_keys.add((sc, pr.strip()))

                # 冲突处理：overwrite 模式下，先删除这些客户的旧规则
                codes_to_update = {r["code"] for r in rules}
                if conflict_mode == "overwrite":
                    filtered_rules = []
                    for r in all_rules:
                        if r.rule_type == "station" and r.stations and r.stations.strip():
                            station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                            if any(sc in codes_to_update for sc in station_list):
                                continue  # 删除这些规则，后面用新的替换
                        filtered_rules.append(r)
                    all_rules = filtered_rules

                # 添加新规则
                for r_data in rules:
                    key = (r_data["code"], r_data["province"])
                    if key in existing_rule_keys and conflict_mode != "overwrite":
                        stats["skipped_rules"] += 1
                        continue
                    new_rule = Rule(
                        name=f"{r_data['code']} - {r_data['group']}",
                        regions=r_data["province"], stations=r_data["code"],
                        min_weight=0.0, max_weight=999.0,
                        first_fee=float(r_data["first_fee"]),
                        continued_fee=float(r_data["continued_fee"]),
                        min_fee=float(r_data["min_fee"]),
                        rule_type="station",
                        continued_unit=r_data.get("continued_unit", "kg"),
                        weight_rounding=r_data.get("weight_rounding", "actual"),
                        rounding_params={},
                    )
                    all_rules.append(new_rule)
                    stats["inserted_rules"] += 1

                self.save_rules(all_rules)
                session.close()
                return {
                    "success": True,
                    "message": f"成功导入 {stats['inserted_customers']} 个客户，{stats['inserted_rules']} 条规则",
                    "stats": stats,
                }
            except Exception as e:
                session.rollback()
                session.close()
                return {"success": False, "message": f"保存失败: {e}", "stats": stats}
        except Exception as e:
            return {"success": False, "message": f"导入失败: {e}",
                    "stats": {"inserted_customers": 0, "inserted_rules": 0}}

    def load_promotion_rules(self) -> List[Dict]:
        """加载活动加价规则"""
        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data.get("promotion_rules", [])
        except Exception:
            return []

    def calculate_fee(self, weight: float, region: str = "", station_code: str = "") -> Dict:
        """
        根据重量、区域、网点计算运费（三层规则匹配 + 活动加价）
        :param weight: 重量(kg)
        :param region: 区域地址
        :param station_code: 客户编码
        :return: {"fee": 金额, "rule_name": 命中的规则名, "is_exception": 是否异常, "remark": 备注}
        """
        # 无重量：使用"无重量默认价格"（不再是异常）
        if weight is None or weight <= 0:
            return {
                "fee": self._empty_weight_fee,
                "rule_name": "无重量默认价",
                "is_exception": False,
                "remark": f"订单无重量，按默认价 ¥{self._empty_weight_fee:.2f} 结算"
            }

        rules = self.load_rules()

        # 第一步：查找客户专属规则（最优先）
        # 修复：stations 字段可能是逗号分隔值，必须按列表匹配；同时在循环内检查区域+重量
        station_rule = None
        sc = (station_code or "").strip()
        sc_list = [s.strip() for s in sc.split(",") if s.strip()] if sc else []
        for r in rules:
            if r.rule_type == "station" and r.stations and r.stations.strip():
                r_station_list = [s.strip() for s in r.stations.split(",") if s.strip()]
                customer_match = any(s in r_station_list for s in sc_list) or (not sc_list and False)
                if customer_match:
                    if r.first_fee > 0 or r.continued_fee > 0 or r.min_fee > 0 or r.regions:
                        # 在循环内直接检查区域+重量，找到完全匹配的客户规则
                        if r.matches(region, station_code, weight):
                            station_rule = r
                            break

        # 如果客户有专属规则且匹配，则使用
        if station_rule:
            matched_rule = station_rule
        else:
            # 第二步：匹配区域规则
            matched_rule = None
            for r in rules:
                if r.rule_type == "region" and r.matches(region, station_code, weight):
                    matched_rule = r
                    break

            # 第三步：使用全局规则兜底
            if not matched_rule:
                for r in rules:
                    if r.rule_type == "global" and r.matches(region, station_code, weight):
                        matched_rule = r
                        break

        if not matched_rule:
            return {
                "fee": 0.0,
                "rule_name": "无匹配规则",
                "is_exception": True,
                "remark": f"区域[{region}] 客户[{station_code}] 未匹配到计费规则"
            }

        # 计算基础运费
        first_fee = matched_rule.first_fee
        continued_fee = matched_rule.continued_fee
        first_weight = 1.0
        continued_unit = matched_rule.continued_unit
        weight_rounding = matched_rule.weight_rounding
        rounding_params = matched_rule.rounding_params

        rounded_weight = apply_weight_rounding(weight, weight_rounding, rounding_params)

        if rounded_weight <= first_weight:
            fee = first_fee
        else:
            continued_weight = rounded_weight - first_weight
            if continued_unit == "100g":
                units = math.ceil(continued_weight / 0.1)
                fee = first_fee + units * continued_fee
            else:
                fee = first_fee + continued_weight * continued_fee

        fee = max(fee, matched_rule.min_fee)
        fee = round(fee, 2)
        base_fee = fee

        # 第四步：应用活动加价（使用统一的日期解析，支持多种格式）
        promo_name = ""
        promo_amount = 0.0
        try:
            from datetime import datetime
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

            # 统一日期解析函数：支持 2026-06-19 / 2026/6/19 / 2026.6.19 / 20260619 / 2026年6月19日
            def _parse_date_local(date_str: str):
                if not date_str:
                    return None
                s = str(date_str).strip()
                if not s:
                    return None
                for ch in ["年", "月", "日", ".", "-", "/"]:
                    s = s.replace(ch, "-")
                while "--" in s:
                    s = s.replace("--", "-")
                s = s.strip("-")
                parts = s.split("-")
                if len(parts) >= 3:
                    try:
                        return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                    except (ValueError, TypeError):
                        pass
                if s.isdigit() and len(s) == 8:
                    try:
                        return datetime(int(s[:4]), int(s[4:6]), int(s[6:8]))
                    except (ValueError, TypeError):
                        pass
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
                    try:
                        return datetime.strptime(str(date_str).strip(), fmt)
                    except Exception:
                        continue
                return None

            promo_rules = self.load_promotion_rules()
            for pr in promo_rules:
                try:
                    start = _parse_date_local(pr.get("start_date", ""))
                    end = _parse_date_local(pr.get("end_date", ""))
                    if start is None or end is None:
                        continue
                    if not (start <= today <= end):
                        continue

                    # 省份限定检查（regions留空=不限定）
                    regions_val = str(pr.get("regions", "")).strip()
                    if regions_val and region:
                        region_kws = [k.strip() for k in regions_val.split(",") if k.strip()]
                        if region_kws and not any(k in region for k in region_kws):
                            continue

                    markup_type = str(pr.get("markup_type", "percent")).strip().lower()
                    try:
                        markup_value = float(str(pr.get("markup_value", "0")).strip())
                    except (ValueError, TypeError):
                        continue

                    if markup_value <= 0:
                        continue

                    if markup_type == "fixed":
                        promo_amount = markup_value
                    elif markup_type == "weight":
                        promo_amount = weight * markup_value
                    elif markup_type == "percent":
                        promo_amount = base_fee * (markup_value / 100.0)
                    else:
                        continue

                    promo_name = str(pr.get("name", "活动加价"))
                    if regions_val:
                        promo_name = f"{promo_name}[{regions_val}]"
                    promo_amount = round(promo_amount, 2)
                    break
                except Exception:
                    continue
        except Exception:
            pass

        if promo_amount > 0:
            fee = round(base_fee + promo_amount, 2)
            return {
                "fee": fee,
                "rule_name": f"{matched_rule.name} + {promo_name}(+¥{promo_amount})",
                "is_exception": False,
                "remark": f"基础运费¥{base_fee}，活动加价¥{promo_amount}，规则类型: {matched_rule.rule_type}"
            }

        return {
            "fee": fee,
            "rule_name": matched_rule.name,
            "is_exception": False,
            "remark": f"规则类型: {matched_rule.rule_type}"
        }

    def get_rules_by_type(self, rule_type: str) -> List[Rule]:
        """按类型获取规则"""
        rules = self.load_rules()
        return [r for r in rules if r.rule_type == rule_type]

    def add_station_rule(self, station_code: str, station_name: str, 
                         first_fee: float, continued_fee: float, min_fee: float):
        """快速添加网点专属规则"""
        rules = self.load_rules()
        new_rule = Rule(
            name=f"{station_name} - 网点专属",
            regions="",
            stations=station_code,
            first_fee=first_fee,
            continued_fee=continued_fee,
            min_fee=min_fee,
            rule_type="station"
        )
        rules.append(new_rule)
        return self.save_rules(rules)

    def copy_region_rules_to_stations(self, station_codes: List[str], region_name: str):
        """将区域规则批量应用到多个网点"""
        rules = self.load_rules()
        
        region_rule = None
        for r in rules:
            if r.rule_type == "region" and region_name in r.name:
                region_rule = r
                break
        
        if not region_rule:
            return False, f"未找到区域规则: {region_name}"
        
        for code in station_codes:
            new_rule = Rule(
                name=f"网点{code} - 继承{region_name}",
                regions="",
                stations=code,
                min_weight=region_rule.min_weight,
                max_weight=region_rule.max_weight,
                first_fee=region_rule.first_fee,
                continued_fee=region_rule.continued_fee,
                min_fee=region_rule.min_fee,
                rule_type="station"
            )
            rules.append(new_rule)
        
        success = self.save_rules(rules)
        return success, f"已为 {len(station_codes)} 个网点创建继承规则"